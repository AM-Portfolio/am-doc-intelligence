import msoffcrypto
import io
import openpyxl
import pandas as pd
import sys

def decrypt_and_read(file_path, password):
    print(f"Decrypting {file_path} with password {password}")
    decrypted = io.BytesIO()
    with open(file_path, "rb") as f:
        file = msoffcrypto.OfficeFile(f)
        file.load_key(password=password)
        file.decrypt(decrypted)
    
    print("Decrypted successfully. Attempting to read as Excel.")
    try:
        wb = openpyxl.load_workbook(decrypted, data_only=True)
        for sheet_name in wb.sheetnames:
            print(f"\n--- Sheet: {sheet_name} ---")
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(min_row=1, max_row=20, values_only=True):
                print(row)
    except Exception as e:
        print(f"openpyxl failed: {e}. Trying pandas with xlrd...")
        try:
            decrypted.seek(0)
            xls = pd.ExcelFile(decrypted, engine="xlrd")
            for sheet in xls.sheet_names:
                print(f"\n--- Sheet: {sheet} ---")
                df = pd.read_excel(xls, sheet_name=sheet, nrows=20, header=None)
                for i, row in df.iterrows():
                    print(f"Row {i}: {row.tolist()}")
        except Exception as e2:
             print(f"Pandas xlrd failed: {e2}")

if __name__ == "__main__":
    decrypt_and_read(sys.argv[1], sys.argv[2])
