import openpyxl

wb = openpyxl.load_workbook("docs/tradebook-ZC3978-FNO_Zerodha.xlsx", read_only=True)
print("Sheets:", wb.sheetnames)

sheet = wb[wb.sheetnames[0]]
for i in range(1, 30):
    row_vals = [sheet.cell(row=i, column=j).value for j in range(1, 15)]
    if any(row_vals):
        print(f"Row {i}: {row_vals}")
